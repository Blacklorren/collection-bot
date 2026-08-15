"""Construit le manifest joueurs de la saison 2 a partir du xlsx roster. (OFFLINE, dev only.)

Source : <xlsx> onglet "Starligue27" -> colonnes A/B/C = Nom / Prenom / Equipe
Sortie : data/roster_s2.json

Chaque joueur recoit un id slug stable (<nom-de-famille>-<prenom>) : c'est la clef
de tout le pipeline images (refs photo -> prompts -> selection -> cutouts). Le club
n'entre PAS dans l'id, pour qu'un transfert ne casse pas les fichiers deja produits.

Le script est idempotent : relancer conserve les champs deja renseignes (poste,
rarete, ref_url, image choisie), ajoute les nouveaux joueurs et marque les partants
"sorti": true au lieu de les supprimer.

Pre-requis : pip install openpyxl

Usage :
    python tools/build_manifest_s2.py
    python tools/build_manifest_s2.py --xlsx "C:/chemin/roster.xlsx" --sheet Starligue27
"""
import argparse
import json
import os
import re
import unicodedata

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(ROOT, "data", "roster_s2.json")
COMPLEMENTS = os.path.join(ROOT, "data", "roster_complements.json")

DEFAULT_XLSX = os.path.join(os.path.expanduser("~"), "Downloads", "Starligue 26.xlsx")
DEFAULT_SHEET = "Starligue27"

# Libelle club canonique (accentue, comme dans cards.json S1) indexe par forme
# repliee ascii/minuscule. slugify() du renderer replie les accents, donc les
# logos assets/logos/<slug>.png restent trouves.
# La rarete n'est pas ecrite dans une colonne : elle est portee par la COULEUR DE
# FOND de la ligne. Les libelles produits sont ceux attendus par RARITY_RGB dans
# utils/card_renderer.py ("Peu Commun" et non "Peu commune"), sinon le degrade de
# rarete n'est pas trouve au rendu.
RARETE_PAR_COULEUR = {
    (241, 194, 50): "Légendaire",    # F1C232 jaune
    (142, 124, 195): "Épique",       # 8E7CC3 violet
    (109, 158, 235): "Rare",         # 6D9EEB bleu
    (147, 196, 125): "Peu Commun",   # 93C47D vert
}
RARETE_SANS_FOND = "Commun"          # cellule non remplie
TOLERANCE = 60                       # distance RGB max pour accepter une nuance voisine

CLUB_CANON = {
    "aix": "Aix",
    "caen": "Caen",
    "cesson-rennes": "Cesson-Rennes",
    "chambery": "Chambéry",
    "chartres": "Chartres",
    "dunkerque": "Dunkerque",
    "limoges": "Limoges",
    "montpellier": "Montpellier",
    "nantes": "Nantes",
    "nimes": "Nîmes",
    "paris": "Paris",
    "saint-raphael": "Saint-Raphaël",
    "saran": "Saran",
    "selestat": "Sélestat",
    "toulouse": "Toulouse",
    "tremblay": "Tremblay",
}


def fold(s):
    """Replie en ascii minuscule sans ponctuation : sert de clef de rapprochement."""
    s = unicodedata.normalize("NFKD", str(s or "")).encode("ascii", "ignore").decode()
    s = re.sub(r"[^A-Za-z0-9]+", "-", s).strip("-").lower()
    return s


def canon_club(raw):
    key = fold(raw)
    return CLUB_CANON.get(key, str(raw or "").strip())


def rarete_from_fill(cell):
    """-> (rarete, rgb_non_reconnu). Cellule sans remplissage = Commun."""
    fill = cell.fill
    rgb = getattr(fill.start_color, "rgb", None)
    if fill.patternType is None or not isinstance(rgb, str) or rgb in ("00000000", "FFFFFFFF"):
        return RARETE_SANS_FOND, None
    try:
        r, g, b = (int(rgb[-6:][i:i + 2], 16) for i in (0, 2, 4))
    except ValueError:
        return None, rgb
    best, dist = None, None
    for (cr, cg, cb), lib in RARETE_PAR_COULEUR.items():
        d = ((r - cr) ** 2 + (g - cg) ** 2 + (b - cb) ** 2) ** 0.5
        if dist is None or d < dist:
            best, dist = lib, d
    return (best, None) if dist <= TOLERANCE else (None, rgb)


def read_rows(xlsx, sheet):
    # data_only=True pour les valeurs, mais les styles (donc les fonds) restent lus
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    if sheet not in wb.sheetnames:
        raise SystemExit(f"Onglet '{sheet}' absent. Onglets dispos : {wb.sheetnames}")
    ws = wb[sheet]
    rows, inconnues = [], {}
    for row in ws.iter_rows(min_row=2, max_col=3):
        nom = (row[0].value or "").strip() if row[0].value else ""
        prenom = (row[1].value or "").strip() if row[1].value else ""
        equipe = (row[2].value or "").strip() if row[2].value else ""
        if not nom or not equipe:
            continue
        rarete, inconnue = rarete_from_fill(row[0])
        if inconnue:
            inconnues.setdefault(inconnue, []).append(f"{prenom} {nom}")
        rows.append({
            "nom_famille": nom,
            "prenom": prenom,
            "club": canon_club(equipe),
            "rarete": rarete,
        })
    return rows, inconnues


def build(rows):
    """Assemble le manifest, id unique (suffixe -2, -3... en cas d'homonymes)."""
    seen = {}
    out = []
    for r in rows:
        base = f"{fold(r['nom_famille'])}-{fold(r['prenom'])}".strip("-")
        seen[base] = seen.get(base, 0) + 1
        pid = base if seen[base] == 1 else f"{base}-{seen[base]}"
        out.append({
            "id": pid,
            # Format d'affichage identique a cards.json S1 : "Gustaf BANKE"
            "nom": f"{r['prenom']} {r['nom_famille']}".strip(),
            "nom_famille": r["nom_famille"],
            "prenom": r["prenom"],
            "club": r["club"],
            "rarete": r["rarete"],
            "poste": None,
            "numero": None,
            "lnh_slug": None,
            "ref_url": None,      # portrait officiel (rempli par fetch_lnh_s2.py)
            "ref_file": None,     # refs/<id>.jpg une fois telecharge
            "image_file": None,   # rendu Midjourney retenu (rempli par le picker)
            "sorti": False,
        })
    return out


# Champs que l'on ne veut jamais ecraser lors d'un rerun
PRESERVE = ("rarete", "poste", "numero", "lnh_slug", "ref_url", "ref_file", "image_file")


def merge(fresh, old_path):
    """Reporte sur le manifest neuf les champs deja renseignes dans l'ancien."""
    if not os.path.exists(old_path):
        return fresh, 0, []
    old = {p["id"]: p for p in json.load(open(old_path, encoding="utf-8"))}
    kept = 0
    for p in fresh:
        prev = old.get(p["id"])
        if not prev:
            continue
        for f in PRESERVE:
            if p.get(f) in (None, "") and prev.get(f) not in (None, ""):
                p[f] = prev[f]
                kept += 1
    # joueurs presents avant et absents du nouveau roster -> conserves, marques sortis
    ids = {p["id"] for p in fresh}
    partis = []
    for pid, prev in old.items():
        if pid in ids:
            continue
        prev["sorti"] = True
        fresh.append(prev)
        partis.append((pid, prev.get("club")))
    return fresh, kept, partis


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default=DEFAULT_XLSX)
    ap.add_argument("--sheet", default=DEFAULT_SHEET)
    args = ap.parse_args()

    if not os.path.exists(args.xlsx):
        raise SystemExit(f"Introuvable : {args.xlsx}")

    rows, inconnues = read_rows(args.xlsx, args.sheet)

    # Joueurs retenus pour la S2 mais absents du xlsx (transferts tardifs, oublis).
    # Sans ce complement, un rebuild les marquerait "sortis" et perdrait leur rarete.
    ajouts = []
    if os.path.exists(COMPLEMENTS):
        comp = json.load(open(COMPLEMENTS, encoding="utf-8"))
        for a in comp.get("ajouts", []):
            ajouts.append({
                "nom_famille": a["nom_famille"],
                "prenom": a.get("prenom", ""),
                "club": canon_club(a["club"]),
                "rarete": a.get("rarete"),
            })
    rows += ajouts

    players = build(rows)
    players, kept, partis = merge(players, OUT)

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    json.dump(players, open(OUT, "w", encoding="utf-8"), ensure_ascii=False, indent=2)

    actifs = [p for p in players if not p.get("sorti")]
    clubs = {}
    for p in actifs:
        clubs[p["club"]] = clubs.get(p["club"], 0) + 1

    print(f"{len(actifs)} joueurs -> {OUT}")
    if ajouts:
        print(f"  dont {len(ajouts)} hors xlsx (data/roster_complements.json) : "
              + ", ".join(f"{a['prenom']} {a['nom_famille']}" for a in ajouts))
    if kept:
        print(f"  {kept} champs deja renseignes conserves du run precedent")
    if partis:
        print(f"  {len(partis)} joueurs sortis de l'effectif (conserves, sorti=true) :")
        for pid, club in partis:
            print(f"    - {pid} ({club})")

    print(f"\n{len(clubs)} clubs :")
    for club in sorted(clubs):
        print(f"  {club:16s} {clubs[club]:3d}")

    # Controle : un club sans logo rendra des cartes sans ecusson
    logos = os.path.join(ROOT, "assets", "logos")
    manquants = [c for c in sorted(clubs) if not os.path.exists(os.path.join(logos, fold(c) + ".png"))]
    if manquants:
        print(f"\nATTENTION logos manquants dans assets/logos/ : {', '.join(manquants)}")
        print("  -> ces cartes seront rendues sans ecusson par compose_v2()")

    # Raretes lues sur la couleur de fond des lignes du xlsx
    ordre = ["Commun", "Peu Commun", "Rare", "Épique", "Légendaire"]
    par_rarete = {}
    for p in actifs:
        par_rarete[p.get("rarete")] = par_rarete.get(p.get("rarete"), 0) + 1
    print("\nRaretes (lues sur la couleur de fond) :")
    for lib in ordre:
        if lib in par_rarete:
            print(f"  {lib:12s} {par_rarete[lib]:3d}")

    if inconnues:
        print("\nATTENTION couleurs de fond non reconnues :")
        for rgb, noms in inconnues.items():
            apercu = ", ".join(noms[:3]) + (" ..." if len(noms) > 3 else "")
            print(f"  #{rgb[-6:]}  {len(noms):3d} joueurs  ({apercu})")
        print("  -> ajoute la teinte dans RARETE_PAR_COULEUR en haut de ce fichier")

    sans_rarete = [p for p in actifs if not p.get("rarete")]
    if sans_rarete:
        print(f"\n{len(sans_rarete)} joueurs sans rarete exploitable.")


if __name__ == "__main__":
    main()
