"""
Ré-injecte les postes remplis dans cards.json.

Lit `postes_a_remplir.xlsx` (colonnes id / poste) et écrit le champ `poste`
sur chaque carte correspondante de cards.json (match par `id`).

- Les cartes Noël (absentes du fichier) restent inchangées (pas de `poste`).
- Sauvegarde l'ancien cards.json en cards.json.bak avant d'écrire.
- Signale les postes manquants ou invalides.

Usage :  py -3 tools/inject_postes.py
"""
import json
import os
import shutil

from openpyxl import load_workbook

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CARDS_PATH = os.path.join(ROOT, "cards.json")
XLSX_PATH = os.path.join(ROOT, "postes_a_remplir.xlsx")

VALID = {"GB", "ALG", "ARG", "DC", "PIV", "ARD", "ALD"}


def main():
    wb = load_workbook(XLSX_PATH, data_only=True)
    ws = wb["Postes"]

    # Repérer les colonnes id et poste depuis l'en-tête
    headers = {str(c.value).strip().lower(): i for i, c in enumerate(ws[1]) if c.value}
    id_col, poste_col = headers.get("id"), headers.get("poste")
    if id_col is None or poste_col is None:
        raise SystemExit("Colonnes 'id' et 'poste' introuvables dans l'onglet Postes.")

    postes_by_id = {}
    manquants, invalides = [], []
    for row in ws.iter_rows(min_row=2):
        cid = row[id_col].value
        if cid is None:
            continue
        poste = row[poste_col].value
        poste = str(poste).strip().upper() if poste is not None else ""
        if not poste:
            manquants.append(cid)
            continue
        if poste not in VALID:
            invalides.append((cid, poste))
            continue
        postes_by_id[cid] = poste

    with open(CARDS_PATH, "r", encoding="utf-8") as f:
        cards = json.load(f)

    updated = 0
    for card in cards:
        if card.get("rarete") == "Noël":
            continue
        if card.get("id") in postes_by_id:
            card["poste"] = postes_by_id[card["id"]]
            updated += 1

    if manquants:
        print(f"⚠️  {len(manquants)} carte(s) sans poste : {manquants[:20]}{' ...' if len(manquants) > 20 else ''}")
    if invalides:
        print(f"❌ {len(invalides)} poste(s) invalide(s) : {invalides[:20]}")

    shutil.copyfile(CARDS_PATH, CARDS_PATH + ".bak")
    with open(CARDS_PATH, "w", encoding="utf-8") as f:
        json.dump(cards, f, ensure_ascii=False, indent=2)

    print(f"✅ {updated} cartes mises à jour. Sauvegarde : cards.json.bak")
    if manquants or invalides:
        print("ℹ️  Complète/corrige le xlsx puis relance ce script.")


if __name__ == "__main__":
    main()
