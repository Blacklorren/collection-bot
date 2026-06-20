"""
Génère `postes_a_remplir.xlsx` à partir de cards.json.

- Une ligne par carte JOUABLE (les cartes Noël sont exclues : promo uniquement).
- Colonne `poste` avec liste déroulante IMPOSÉE (validation stricte) :
  GB, ALG, ARG, DC, PIV, ARD, ALD.
- Trié par club puis nom pour remplir vite.

Usage :  py -3 tools/generate_postes_xlsx.py
"""
import json
import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CARDS_PATH = os.path.join(ROOT, "cards.json")
OUT_PATH = os.path.join(ROOT, "postes_a_remplir.xlsx")

POSTES = ["GB", "ALG", "ARG", "DC", "PIV", "ARD", "ALD"]
POSTE_LABELS = {
    "GB": "Gardien de but",
    "ALG": "Ailier gauche",
    "ARG": "Arrière gauche",
    "DC": "Demi-centre",
    "PIV": "Pivot",
    "ARD": "Arrière droit",
    "ALD": "Ailier droit",
}
HEADERS = ["id", "nom", "club", "rarete", "poste"]


def main():
    with open(CARDS_PATH, "r", encoding="utf-8") as f:
        cards = json.load(f)

    # Cartes jouables uniquement (Noël = promo, exclue du jeu)
    playable = [c for c in cards if c.get("rarete") != "Noël"]
    playable.sort(key=lambda c: (str(c.get("club", "")).lower(), str(c.get("nom", "")).lower()))

    wb = Workbook()
    ws = wb.active
    ws.title = "Postes"

    header_fill = PatternFill("solid", fgColor="2F4858")
    header_font = Font(bold=True, color="FFFFFF")
    poste_fill = PatternFill("solid", fgColor="FFF2CC")  # colonne à remplir en surbrillance
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, name in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for i, card in enumerate(playable, start=2):
        ws.cell(row=i, column=1, value=card.get("id")).border = border
        ws.cell(row=i, column=2, value=card.get("nom")).border = border
        ws.cell(row=i, column=3, value=card.get("club")).border = border
        ws.cell(row=i, column=4, value=card.get("rarete")).border = border
        pcell = ws.cell(row=i, column=5)
        pcell.fill = poste_fill
        pcell.border = border
        pcell.alignment = Alignment(horizontal="center")

    last_row = len(playable) + 1

    # Liste déroulante IMPOSÉE sur la colonne poste (E2:E{last})
    dv = DataValidation(
        type="list",
        formula1='"%s"' % ",".join(POSTES),
        allow_blank=True,
        showErrorMessage=True,
    )
    dv.error = "Poste invalide. Choisis dans la liste : " + ", ".join(POSTES)
    dv.errorTitle = "Poste non autorisé"
    dv.prompt = "Choisis un poste dans la liste déroulante"
    dv.promptTitle = "Poste"
    ws.add_data_validation(dv)
    dv.add("E2:E%d" % last_row)

    # Largeurs de colonnes
    widths = {"A": 6, "B": 30, "C": 22, "D": 14, "E": 10}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.freeze_panes = "A2"

    # Onglet légende des postes
    legend = wb.create_sheet("Légende")
    legend.cell(row=1, column=1, value="Code").font = Font(bold=True)
    legend.cell(row=1, column=2, value="Poste").font = Font(bold=True)
    for r, code in enumerate(POSTES, start=2):
        legend.cell(row=r, column=1, value=code)
        legend.cell(row=r, column=2, value=POSTE_LABELS[code])
    legend.column_dimensions["A"].width = 8
    legend.column_dimensions["B"].width = 20

    wb.save(OUT_PATH)
    print(f"OK : {OUT_PATH}  ({len(playable)} cartes jouables, Noël exclues)")


if __name__ == "__main__":
    main()
